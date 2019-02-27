import openpyxl

import argparse
import logging
import re

logger = logging.getLogger()
logger.setLevel(logging.INFO)

INSTRUCTOR_NAME_REGEX = re.compile("^(.*) Class\/Enrollments.*$")


def row_is_blank(row):
    """Returns true iff the row is blank"""
    for cell in row:
        # Skip merged cells, which do not have a value.
        try:
            if cell.value is not None:
                return False
        except AttributeError:
            continue

    # We have iterated over all cells in the row and none had a value.
    return True

def copy_row(dest_sheet, row):
    """Deep copies the cell values into a new row in the dest sheet

    openpyxl has an API limit that prevents direct copying of the cell
    objects from one workbook to another.

    """
    new_row = []
    for cell in row:
        # Skip merged cells, which do not have a value.
        try:
            new_row.append(cell.value)
        except AttributeError:
            # Insert empty cells if the cells were originally merged
            new_row.append(None)

    dest_sheet.append(new_row)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
    description='Split an xlsx mindbody payroll report into individual payroll reports.')

    parser.add_argument('source_file',
                        metavar='source_file',
                        type=str,
                        help="a source mindbody payroll xlsx file")

    parser.add_argument('-p', '--pay_period',
                        type=str,
                        help="The pay period string to append to each created filename")

    args = parser.parse_args()

    if args.pay_period is None:
        logging.critical('pay_period is a required parameter, try --pay_period "date"')
        exit(1)

    # Load the excel sheet file into python
    source_workbook = openpyxl.load_workbook(args.source_file)

    # Access the first sheet in the xlsx file.
    #
    # Note that in the future if mindbody changes their format, this
    # may also need to change.
    source_sheet = source_workbook[source_workbook.sheetnames[0]]

    #loop over all rows in the source sheet.
    #find blank rows, after each blank row...
    #get the name, which will be the next row, first cell
    #create a new book with the new name
    #copy rows into the new book
    #when find blank row, save the book, start over.
    dest_workbook = None
    dest_sheet = None
    current_name = None
    for row_number, row in enumerate(source_sheet.rows, 1):
        logging.debug("currently processing source row %d", row_number)
        logging.debug("row Contents: %s", row)

        if row_is_blank(row):
            logging.info("---------------Begin Iteration------------------")
            logging.info("beginning a new split on row %d", row_number)
            # When we find a blank row, begin a new split.
            if dest_workbook is not None and current_name is not None:
                filename = current_name + " - " + args.pay_period + ".xlsx"
                dest_workbook.save(filename)

            dest_workbook = openpyxl.Workbook()
            dest_sheet = dest_workbook.active
            current_name = source_sheet.cell(row_number + 1, 1).value
            logging.debug("current_name: %s", current_name)
            # Clean the current name to extract the actual instructor
            # name. Mindbody mushes the instructor name and other
            # metadata into the same cell.
            if INSTRUCTOR_NAME_REGEX.match(current_name):
                # The python regex api is tricky, be sure to consult
                # the documentation before changing this line.
                current_name = INSTRUCTOR_NAME_REGEX.match(current_name).group(1)
                logging.info("section instructor is %s", current_name)
            else:
                logging.warning(
                    "section has malformed name '%s' and will not be recorded",
                    current_name)
                current_name = None

            logging.debug("new split name: %s", current_name)
        else:
            # When the row is not blank, just copy over!
            copy_row(dest_sheet, row)
